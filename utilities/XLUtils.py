from pathlib import Path
from typing import Any, List
import openpyxl
from openpyxl.styles import PatternFill

# Reusable fill styles
GREEN_FILL = PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def get_row_count(file_path: str | Path, sheet_name: str) -> int:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    workbook.close()
    return total_rows


def get_column_count(file_path: str | Path, sheet_name: str) -> int:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook[sheet_name]
    total_cols = sheet.max_column
    workbook.close()
    return total_cols


def read_data(file_path: str | Path, sheet_name: str, row_num: int, column_no: int) -> Any:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    value = sheet.cell(row=row_num, column=column_no).value
    workbook.close()
    return value


def write_data(file_path: str | Path, sheet_name: str, row_num: int, column_no: int, data: Any) -> None:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_no).value = data
    workbook.save(file_path)
    workbook.close()


def fill_green_color(file_path: str | Path, sheet_name: str, row_num: int, column_no: int) -> None:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_no).fill = GREEN_FILL
    workbook.save(file_path)
    workbook.close()


def fill_red_color(file_path: str | Path, sheet_name: str, row_num: int, column_no: int) -> None:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num, column=column_no).fill = RED_FILL
    workbook.save(file_path)
    workbook.close()


def get_all_rows(file_path: str | Path, sheet_name: str) -> List[tuple]:
    """Helper to fetch all rows efficiently at once."""
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return data
